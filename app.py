import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import zipfile

st.set_page_config(page_title="Procesador de Ventas", layout="wide")
st.title(" Procesador de Reportes de Ventas y Compras")

# Campos requeridos por defecto para VENTAS
CAMPOS_ORIGINALES_VENTAS = [
    "Fecha de emisión",
    "Fecha Vcto/Pago",
    "Tipo CP/Doc.",
    "Serie del CDP",
    "Nro CP o Doc. Nro Inicial (Rango)",  
    "Nro Doc Identidad",
    "Apellidos Nombres/ Razón Social",
    "BI Gravada",
    "IGV / IPM",
    "Total CP",
    "Moneda",
]

# Campos requeridos por defecto para COMPRAS
CAMPOS_ORIGINALES_COMPRAS = [
    "Fecha de emisión",
    "Fecha Vcto/Pago",
    "Tipo CP/Doc.",
    "Serie del CDP",
    "Nro CP o Doc. Nro Inicial (Rango)",
    "Tipo Doc Identidad",
    "Nro Doc Identidad",
    "Apellidos Nombres/ Razón  Social",
    "BI Gravado DG",
    "IGV / IPM DG",
    "Valor Adq. NG",
    "Total CP",
    "Moneda",
    "Fecha Emisión Doc Modificado",
    "Serie CP Modificado",
    "Nro CP Modificado",
]


RENOMBRAR = {
    "Tipo CP/Doc.": "Tipo Doc",
    "Serie del CDP": "Serie",
    "Nro CP o Doc. Nro Inicial (Rango)":"Nro",
}

# Diccionario de RUCs y nombres de empresas
EMPRESAS_RUC = {

    "20600081650": "BALANCEADOS BLANCA & JOSE LUIS E.I.R.L.",
    "10409958970": "SOSA COSTI RODOLFO NELSON",
    "20526165650": "EMPRESA DE TRANSPORTES DE PASAJEROS SANTA SOFIA SOCIEDAD DE RESPONSABILIDAD LIMITADA",
    "10036548610": "MENDOZA CHAVEZ DORIS",
    "20526543271": "EMPRESA DE TRANSPORTES DE PASAJEROS LA QUINTA S.R.L.",
    "10036868371": "CARDOZA JIMENEZ LILIAN VERONICA",
    "20525390561": "ETP MALLARES S.R.L.",
    "20608133811": "EMPRESA DE TRANSPORTES FORTALEZA AMIGOS UNIDOS DE SULLANA S.A.C.",
    "20605393951": "EMPRESA DE TRANSPORTE Y SERVICIOS GENERALES NARCISA DE JESUS E.I.R.L.",
    "10763165821": "REVOLLEDO GUTIERREZ DARWIN ALEXIS",
    "20611795271": "ROSSALUD PHARMACY E.I.R.L.",
    "20526310412": "EMPRESA DE TRANSPORTES DIOS MIO SOCIEDAD COMERCIAL DE RESPONSABILIDAD LIMITADA",
    "10751358194": "MEZA MONTERO HENRY JOEL",
    "10460433946": "ALAMA ABARCA ROSA MERCEDES",
    "10479164474": "UMBO DOMINGUEZ CESAR AUGUSTO",
    "10413501194": "MOGOLLON PRADO CAROLINA DE JESUS",
    "10027617374": "MADRID VARGAS PEDRO TEODOMIRO",
    "10103145274": "GUIZAR FERNANDEZ CARLOS HUMBERTO",
    "20526245335": "EMPRESA DE TRANSPORTES QUERECOTILLO TOURS SOCIEDAD ANONIMA CERRADA",
    "20603114745": "TRANSPORTES MERARI S.A.",
    "10722713295": "LORO FARFAN ANGEE SOPHIA DE FATIMA",
    "10414088495": "CAMPOS BENITES FRANCISCO JAVIER",
    "20609716585": "FERRETERIA & SERVICENTRO EL PARTIDOR-FERRESERVI EMPRESA INDIVIDUAL DE RESPONSABILIDAD LIMITADA",
    "10486467555": "GRANDA SERNAQUE HENRRY DAVID",
    "10751342115": "ZAPATA ESPIL MAURICIO JUNIOR",
    "10028979385": "CARLIN RUIZ EDUARDO",
    "10432666706": "CARRASCO CASTRO JUAN MANUEL",
    "10036653316": "LADINES GONZALES GUSTAVO",
    "10434042696": "ARGANDOÑA SINARAHUA JESSENIA DEL PILAR",
    "20612032476": "SERVICIOS & NEGOCIOS GENERALES GUADMER S.A.C",
    "20601019176": "COMERCIAL ALEXANDER C & V E.I.R.L.",
    "10473696687": "LUPUCHE NAVARRO CARLOS ALBERTO",
    "10036736327": "CARDOZA JIMENEZ YSELA CAROLINA",
    "20526264127": "EMPRESA TRANSPORTE MALLASULL S.A.",
    "10413153447": "PAULINI HUANCA GILBERTO",
    "20530205348": "TRANSPORTES SEGUNDO EDUARDO REQUEJO SOCIEDAD ANONIMA CERRADA - TRANSEDUR S.A.C",
    "20612709808": "COMERCIAL UMBO E.I.R.L",
    "10435850508": "GARCIA MONTERO JULIO CESAR",
    "10453609389": "GARCIA MONTERO MARIA DEL PILAR",
    "20603268629": "R & E PREVENTION SERVICIOS GENERALES S.A.C.", 
}

def extraer_ruc_de_nombre_archivo(filename):
    """
    Extrae el RUC del nombre del archivo.
    Busca un patrón de 11 dígitos en el nombre del archivo.
    """
    import re
    # Buscar secuencia de exactamente 11 dígitos
    patron = r'(\d{11})'
    matches = re.findall(patron, filename)
    
    # Retornar el primer RUC válido encontrado (que empiece con 10 o 20)
    for match in matches:
        if match.startswith('10') or match.startswith('20'):
            return match
    
    return None


opcion = st.radio("Selecciona el tipo de archivo que deseas procesar:", ["Ventas", "Compras"], horizontal=True)



def read_file(f):
    if f.name.lower().endswith(".zip"):
        with zipfile.ZipFile(f) as z:
    
            csv_files = [name for name in z.namelist() if name.lower().endswith(".csv")]
            if not csv_files:
                st.error("El archivo ZIP no contiene archivos CSV.")
                return None
            with z.open(csv_files[0]) as csv_file:
                try:
                
                    return pd.read_csv(csv_file, encoding="utf-8", engine="python")
                except pd.errors.ParserError as e:
                    st.warning(f"Error de formato detectado: {str(e)}")
                    st.info("🔧 Intentando leer con configuración alternativa...")
                    csv_file.seek(0)
                    return pd.read_csv(csv_file, encoding="utf-8", engine="python", 
                                    on_bad_lines='skip', sep=',', quotechar='"')
    elif f.name.lower().endswith(".csv"):
        try:
            return pd.read_csv(f, encoding="utf-8", engine="python")
        except pd.errors.ParserError as e:
            st.warning(f"Error de formato detectado: {str(e)}")
            st.info("🔧 Intentando leer con configuración alternativa...")
            f.seek(0)
            return pd.read_csv(f, encoding="utf-8", engine="python", 
                            on_bad_lines='skip', sep=',', quotechar='"')
    return pd.read_excel(f)

def clean_numeric_series(s):
    return pd.to_numeric(s.astype(str).str.replace(r'[^\d\.\-]', '', regex=True), errors="coerce")

def detect_missing_correlatives(df, serie_col="Serie", numero_col="Nro"):
    """
    Detecta números correlativos faltantes en las boletas por serie
    """
    missing_report = []
    
    for serie in df[serie_col].unique():
        serie_data = df[df[serie_col] == serie].copy()

        try:
            numeros = pd.to_numeric(serie_data[numero_col], errors='coerce').dropna().astype(int)
            numeros = sorted(numeros.unique())
            
            if len(numeros) > 1:

                min_num = min(numeros)
                max_num = max(numeros)

                secuencia_completa = set(range(min_num, max_num + 1))
                numeros_existentes = set(numeros)
                
                faltantes = sorted(secuencia_completa - numeros_existentes)
                
                if faltantes:
                    missing_report.append({
                        'Serie': serie,
                        'Rango': f"{min_num}-{max_num}",
                        'Faltantes': faltantes,
                        'Total_Faltantes': len(faltantes)
                    })
        except:
            continue
    
    return missing_report

def get_dynamic_title(df_, base_title, filename=None):
    """
    Genera un título dinámico basado en la fecha de emisión del DataFrame
    y el RUC extraído del nombre del archivo
    """
    meses = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    
    # Extraer nombre de empresa del RUC en el nombre del archivo
    nombre_empresa = ""
    if filename:
        ruc = extraer_ruc_de_nombre_archivo(filename)
        if ruc and ruc in EMPRESAS_RUC:
            nombre_empresa = f" - {EMPRESAS_RUC[ruc]}"
    
    try:
        # Buscar la columna de fecha de emisión
        if "Fecha de emisión" in df_.columns:
            # Obtener la primera fecha válida (ignorar valores vacíos o texto como "FACTURAS", "BOLETAS")
            for fecha in df_["Fecha de emisión"]:
                if pd.notna(fecha) and str(fecha).strip() not in ["", "FACTURAS", "BOLETAS"]:
                    # Intentar convertir a datetime con formato DD/MM/YYYY
                    fecha_dt = pd.to_datetime(fecha, format='%d/%m/%Y', errors='coerce')
                    if pd.notna(fecha_dt):
                        mes = meses.get(fecha_dt.month, "")
                        anio = fecha_dt.year
                        return f"{base_title} - {mes} {anio}{nombre_empresa}"
    except:
        pass
    
    # Si no se puede determinar la fecha, retornar el título base con empresa
    return f"{base_title}{nombre_empresa}" if nombre_empresa else base_title

def to_excel_bytes_with_title(df_, title):
    from openpyxl.styles import Font, Alignment, Border, Side
    
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        
        df_.to_excel(writer, index=False, startrow=1, sheet_name="Reporte")
        worksheet = writer.sheets["Reporte"]
        
    
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_.columns))
        cell = worksheet.cell(row=1, column=1)
        cell.value = title
        cell.font = Font(name="Arial", bold=True, size=8)
        cell.alignment = Alignment(horizontal="center")
        for col_idx, column in enumerate(df_.columns, start=1):
            # Calcular el ancho basándose solo en el contenido de los datos
            max_length = 0
            for row in df_.values:
                if row[col_idx-1] is not None and str(row[col_idx-1]).strip() != "":
                    max_length = max(max_length, len(str(row[col_idx-1])))
            # Ancho mínimo de 8 y máximo de 50
            adjusted_width = max(8, min(max_length + 1, 50))
            worksheet.column_dimensions[worksheet.cell(row=2, column=col_idx).column_letter].width = adjusted_width

            header_cell = worksheet.cell(row=2, column=col_idx)
            header_cell.font = Font(name="Arial", bold=True, size=8)
            header_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=len(df_.columns)):
            for cell in row:
                cell.font = Font(name="Arial", size=8)
        
        # Agregar línea negra encima de las filas de totales y debajo de FACTURAS/BOLETAS
        for row_idx in range(3, worksheet.max_row + 1):
            # Buscar en todas las columnas de la fila
            row_text = ""
            for col_idx in range(1, len(df_.columns) + 1):
                cell_value = str(worksheet.cell(row=row_idx, column=col_idx).value)
                row_text += cell_value.upper() + " "
            
            # Buscar filas que contengan "TOTAL" - línea superior y negrita
            if "TOTAL" in row_text:
                black_border = Border(top=Side(style='thin', color='000000'))
                for col_idx in range(1, len(df_.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = black_border
                    cell.font = Font(name="Arial", bold=True, size=8)
            
            # Buscar filas que contengan "FACTURAS" o "BOLETAS" - línea inferior
            if "FACTURAS" in row_text or "BOLETAS" in row_text:
                black_border = Border(bottom=Side(style='thin', color='000000'))
                for col_idx in range(1, len(df_.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = black_border
        
        # Configuración de página para impresión
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.paperSize = 9  # 9 = A4
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    
    return out.getvalue()

if opcion == "Ventas":
    st.header("Subir archivo de Ventas")
    uploaded_file = st.file_uploader("Sube tu archivo CSV, Excel o ZIP (conteniendo un CSV)", type=["csv", "xlsx", "zip"], key="ventas")

    if uploaded_file:
        df = read_file(uploaded_file)
        
        # Mostrar información de la empresa detectada
        ruc_detectado = extraer_ruc_de_nombre_archivo(uploaded_file.name)
        if ruc_detectado and ruc_detectado in EMPRESAS_RUC:
            st.info(f"🏢 Empresa detectada: {EMPRESAS_RUC[ruc_detectado]} (RUC: {ruc_detectado})")
        else:
            if ruc_detectado:
                st.warning(f"⚠️ RUC detectado ({ruc_detectado}) no está registrado en el sistema. Agréguelo al diccionario EMPRESAS_RUC.")
            else:
                st.warning(f"⚠️ No se detectó ningún RUC en el nombre del archivo: {uploaded_file.name}")
        
        if st.checkbox(" Mostrar vista previa del archivo original"):
            st.subheader(" Vista previa del archivo")
            st.dataframe(df.head(10))
        columnas_existentes = [c for c in CAMPOS_ORIGINALES_VENTAS if c in df.columns]
        faltantes = [c for c in CAMPOS_ORIGINALES_VENTAS if c not in df.columns]

        if faltantes:
            st.warning(f"No se encontraron estas columnas en tu archivo: {faltantes}")
        extra_cols = st.multiselect("Selecciona columnas adicionales (si deseas)", [c for c in df.columns if c not in columnas_existentes])
        df = df[columnas_existentes + extra_cols].copy()
        df = df.rename(columns=RENOMBRAR)
        for c in ["BI Gravada", "IGV / IPM", "Total CP"]:
            if c in df.columns:
                df[c] = clean_numeric_series(df[c])
        
        # Calcular totales excluyendo columnas que no deben sumarse
        exclude_columns = ["Tipo Doc", "Nro", "Nro Doc Identidad", "Fecha de emisión", "Fecha Vcto/Pago"]
        numeric_columns = df.select_dtypes(include=np.number).columns
        columns_to_sum = [col for col in numeric_columns if col not in exclude_columns]
        totals = df[columns_to_sum].sum()
        
        total_row = {col: "" for col in df.columns}
        total_row["Apellidos Nombres/ Razón Social"] = "TOTAL VENTAS"
        for c in totals.index:
            total_row[c] = round(totals[c], 2)

        df_with_total = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        st.subheader("Reporte final con Totales")
        st.dataframe(df_with_total)
        st.subheader("Descargar Reporte con Totales- SIN AGRUPAR")
        title = get_dynamic_title(df_with_total, "REPORTE DE VENTAS", uploaded_file.name)
        xlsx_bytes_totales = to_excel_bytes_with_title(df_with_total, title)
        st.download_button("⬇Descargar Excel con Totales", xlsx_bytes_totales, file_name="reporte_ventas_totales.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="ventas_totales")
        if st.checkbox("Agrupar boletas por fecha"):
            if "Tipo Doc" in df.columns and "Fecha de emisión" in df.columns:
                mask_boleta = df["Tipo Doc"] == 3  
                boletas = df[mask_boleta]
                otros = df[~mask_boleta]  
                if not boletas.empty:
                    missing_correlatives = detect_missing_correlatives(boletas)
                    
                    if missing_correlatives:
                        st.warning("**ADVERTENCIA: Se detectaron números correlativos faltantes en las boletas**")
                        st.subheader(" Números de Boletas Faltantes por Serie")
                        
                        for item in missing_correlatives:
                            st.error(f"**Serie {item['Serie']}** (Rango: {item['Rango']})")
                            st.write(f" **Números faltantes ({item['Total_Faltantes']}):** {', '.join(map(str, item['Faltantes']))}")
                            st.write("---")
                    else:
                        st.success(" **Secuencia correlativa completa - No se detectaron números faltantes**")

                if mask_boleta.any():
                    grouped_boletas = (
                        boletas
                        .groupby([boletas["Fecha de emisión"], "Serie"])
                        .agg({
                            "Nro": lambda x: f"{min(x)}-{max(x)}", 
                            "BI Gravada": lambda x: round(x.sum(), 2),
                            "IGV / IPM": lambda x: round(x.sum(), 2),
                            "Total CP": lambda x: round(x.sum(), 2),
                            "Moneda": lambda x: x.iloc[0] if len(set(x)) == 1 else "VARIAS"
                        })
                        .reset_index()
                    )
                    grouped_boletas["Apellidos Nombres/ Razón Social"] = "CLIENTE VARIOS"
                    grouped_boletas["Tipo Doc"] = 3 
                else:
                    grouped_boletas = pd.DataFrame()

                final_report = pd.concat([
                    otros[otros["Tipo Doc"] == 1],  
                    otros[otros["Tipo Doc"] == 7], 
                    grouped_boletas,  
                    otros[(otros["Tipo Doc"] != 1) & (otros["Tipo Doc"] != 7)] 
                ], ignore_index=True)

                def calculate_totals(df, exclude_columns):
                    numeric_columns = df.select_dtypes(include=np.number).columns
                    columns_to_sum = [col for col in numeric_columns if col not in exclude_columns]
                    return df[columns_to_sum].sum()

                exclude_columns = ["Tipo Doc", "Nro", "Nro Doc Identidad", "Fecha de emisión", "Fecha Vcto/Pago"]

                total_facturas = calculate_totals(final_report[final_report["Tipo Doc"] == 1], exclude_columns)
                total_boletas = calculate_totals(final_report[final_report["Tipo Doc"] == 3], exclude_columns)

                total_general = total_facturas + total_boletas
                total_general_row = {col: "" for col in final_report.columns}
                total_general_row.update({
                    "Apellidos Nombres/ Razón Social": "TOTAL GENERAL VENTAS"
                })
                for col in total_general.index:
                    total_general_row[col] = round(total_general[col], 2)
                total_facturas_row = {col: "" for col in final_report.columns}
                total_facturas_row.update({
                    "Apellidos Nombres/ Razón Social": "TOTAL FACTURAS"
                })
                for col in total_facturas.index:
                    total_facturas_row[col] = round(total_facturas[col], 2)
                total_boletas_row = {col: "" for col in final_report.columns}
                total_boletas_row.update({
                    "Apellidos Nombres/ Razón Social": "TOTAL BOLETAS"
                })
                for col in total_boletas.index:
                    total_boletas_row[col] = round(total_boletas[col], 2)
                facturas = final_report[final_report["Tipo Doc"] == 1]
                boletas = final_report[final_report["Tipo Doc"] == 3]
                otros = final_report[(final_report["Tipo Doc"] != 1) & (final_report["Tipo Doc"] != 3)]
                final_report = pd.concat([
                    pd.DataFrame([{col: "" for col in final_report.columns}]).assign(**{"Fecha de emisión": "FACTURAS"}), 
                    facturas,
                    pd.DataFrame([total_facturas_row]),
                    pd.DataFrame([{col: "" for col in final_report.columns}]).assign(**{"Fecha de emisión": "BOLETAS"}),  
                    boletas,
                    pd.DataFrame([total_boletas_row]),
                    pd.DataFrame([total_general_row]),
                    otros
                ], ignore_index=True)

                st.subheader(" Reporte final con agrupación y totales generales")
                st.dataframe(final_report)

                title = get_dynamic_title(final_report, "REPORTE DE VENTAS", uploaded_file.name)
                xlsx_bytes = to_excel_bytes_with_title(final_report, title)
                st.download_button("⬇Descargar Excel AGRUPADO", xlsx_bytes, file_name="reporte_ventas_agrupado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="ventas_agrupado")

elif opcion == "Compras":
    st.header("Subir archivo de Compras")
    uploaded_file = st.file_uploader("Sube tu archivo CSV, Excel o ZIP (conteniendo un CSV)", type=["csv", "xlsx", "zip"], key="compras")

    if uploaded_file:
        df = read_file(uploaded_file)
        
        # Mostrar información de la empresa detectada
        ruc_detectado = extraer_ruc_de_nombre_archivo(uploaded_file.name)
        if ruc_detectado and ruc_detectado in EMPRESAS_RUC:
            st.info(f"🏢 Empresa detectada: {EMPRESAS_RUC[ruc_detectado]} (RUC: {ruc_detectado})")
        else:
            if ruc_detectado:
                st.warning(f"⚠️ RUC detectado ({ruc_detectado}) no está registrado en el sistema. Agréguelo al diccionario EMPRESAS_RUC.")
            else:
                st.warning(f"⚠️ No se detectó ningún RUC en el nombre del archivo: {uploaded_file.name}")

        if st.checkbox("Mostrar vista previa del archivo"):
            st.subheader("Vista previa del archivo")
            st.dataframe(df.head(10))

        columnas_existentes = [c for c in CAMPOS_ORIGINALES_COMPRAS if c in df.columns]
        faltantes = [c for c in CAMPOS_ORIGINALES_COMPRAS if c not in df.columns]

        if faltantes:
            st.warning(f" No se encontraron estas columnas en tu archivo: {faltantes}")
        extra_cols = st.multiselect("Selecciona columnas adicionales (si deseas)", [c for c in df.columns if c not in columnas_existentes])
        
        cols_to_remove = st.multiselect(" Selecciona columnas que deseas QUITAR (si deseas)", columnas_existentes)
        
        columnas_finales = [c for c in columnas_existentes if c not in cols_to_remove]

        df = df[columnas_finales + extra_cols].copy()

        df = df.rename(columns=RENOMBRAR)
        columnas_numericas_compras = ["BI Gravado DG", "IGV / IPM DG", "Valor Adq. NG", "Total CP"]
        for c in columnas_numericas_compras:
            if c in df.columns:
                df[c] = clean_numeric_series(df[c])

        def calculate_totals_compras(df, exclude_columns):
            numeric_columns = df.select_dtypes(include=np.number).columns
            columns_to_sum = [col for col in numeric_columns if col not in exclude_columns]
            return df[columns_to_sum].sum()

        exclude_columns_compras = ["Tipo Doc", "Nro", "Tipo Doc Identidad", "Nro Doc Identidad", "Nro CP Modificado", "Fecha de emisión", "Fecha Vcto/Pago", "Fecha Emisión Doc Modificado"]
        totals = calculate_totals_compras(df, exclude_columns_compras)

        total_row = {col: "" for col in df.columns}
        total_row["Apellidos Nombres/ Razón  Social"] = "TOTAL COMPRAS"
        for c in totals.index:
            total_row[c] = round(totals[c], 2)

        df_with_total = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        st.subheader("Reporte final con Totales")
        st.dataframe(df_with_total)
        title = get_dynamic_title(df_with_total, "REPORTE DE COMPRAS", uploaded_file.name)
        xlsx_bytes = to_excel_bytes_with_title(df_with_total, title)
        st.download_button(
            "⬇ Descargar Excel final",
            xlsx_bytes,
            file_name="reporte_compras.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="compras_final"
        )